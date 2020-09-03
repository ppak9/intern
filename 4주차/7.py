# 실습과제(숙제) 네이버 실시간 검색어 1등 기사 가져오기

import requests
from bs4 import BeautifulSoup
import openpyxl

# f= open('test.csv','w')
# f.write('제목')
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36'}

test =[]

# User-agent
url = 'https://datalab.naver.com/keyword/realtimeList.naver'
raw = requests.get(url,headers=headers)
soup = BeautifulSoup(raw.content,'html.parser')
result = soup.select('span.item_title_wrap')
wb = openpyxl.Workbook()

for item in result:
    keyword = item.select_one('span.item_title').text.strip()
    test.append(keyword)
# 동시에 네이버 실시간 검색어 1위 기사 가지고 오기
try:
    wb=openpyxl.load_workbook("navernews_xlsx")
    sheet = wb.active
    print("불러오기 완료")
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["제목","언론사"])

for n in range(1,100,10):
    raw = requests.get('https://search.naver.com/search.naver?where=news&query='+test[0])
    soup = BeautifulSoup(raw.text,'html.parser')
    articles = soup.select("ul.type01 > li")

    for ar in articles:
        title = ar.select_one("a._sp_each_title").text
        company = ar.select_one('span._sp_each_source').text
        print(title,company)
        sheet.append([title,company])

wb.save('navertv.xlsx')

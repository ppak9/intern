# 4,5 xlsx로 묶은 이후 마지막 실습으로 변경 
# 마지막 실습으로 변경
import requests
from bs4 import BeautifulSoup
import openpyxl
# from urllib.parse import quote_plus

headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36'}

test =[]

# User-agent
url = 'https://datalab.naver.com/keyword/realtimeList.naver'
raw = requests.get(url,headers=headers)
soup = BeautifulSoup(raw.content,'html.parser')
result = soup.select('span.item_title_wrap')
wb = openpyxl.Workbook()

for item in result:
    keyword = item.select_one('span.item_title').text.strip()
    test.append(keyword)
    # print(type(test[0]))
# 동시에 네이버 실시간 검색어 1위 기사 가지고 오기
# 강의 내용: 예외처리 + load에 대한 설명
try:
    wb=openpyxl.load_workbook("navernews_xlsx")
    sheet = wb.active
    print("불러오기 완료")
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["제목","언론사"])

# 강의 내용: 추가적인 html 구조 파악
for n in range(1,100,10):
    base_url = 'https://search.naver.com/search.naver?where=news&query='
    url = base_url + test[0] + str(n)
    raw = requests.get(url,headers=headers)
    # url을 통해 기사 수집
    soup = BeautifulSoup(raw.text,'html.parser')
    articles = soup.select("ul.type01 > li")

    for ar in articles:
        title = ar.select_one("a._sp_each_title").text
        company = ar.select_one('span._sp_each_source').text
        print(title,company)
        sheet.append([title,company])

wb.save('new_naver.xlsx')

import openpyxl
import requests
from bs4 import BeautifulSoup

wb = openpyxl.Workbook()

sheet = wb.active

#직접 지정
sheet['A1'] ="hello world!"
sheet1.title ="1st sheet"
# 직접 지정 2
sheet2 = wb.create_sheet('2nd sheet')
sheet1.cell(row=3,column=3).value="BYE!"

#마지막 데이터 추가
subject =['박종현','박종민','이민철','권영권']
sheet2.apppend(subject)

for i in range(i):
    sheet.cell(row=i,column=i).value=i

wb.save('test2.xlsx')